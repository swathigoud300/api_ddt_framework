import openpyxl
import json
import requests
import jsonpath
def test_add_multiple_users():
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open("C:\\Users\\jagad\\mul_users.json")
    json_request = json.loads(file.read())

    workbook = openpyxl.load_workbook("C:\\Users\\jagad\\multiple_users.xlsx")
    sheet = workbook['Sheet1']
    rows = sheet.max_row

    for i in range(1,rows+1):
        cell_first_name = sheet.cell(row=i,column=1)
        cell_middle_name = sheet.cell(row=i,column=2)
        cell_last_name = sheet.cell(row=i,column=3)
        cell_date_of_birth = sheet.cell(row=i,column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(api_url,json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201

